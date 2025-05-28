import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

st.title("Krank-Meldungen Monatsübersicht (ab 2025)")

uploaded_files = st.file_uploader("Excel-Dateien hochladen", type=["xlsx"], accept_multiple_files=True)

german_months = {
    1: "Januar", 2: "Februar", 3: "März", 4: "April",
    5: "Mai", 6: "Juni", 7: "Juli", 8: "August",
    9: "September", 10: "Oktober", 11: "November", 12: "Dezember"
}

wochentage = [
    "Montag", "Dienstag", "Mittwoch", "Donnerstag",
    "Freitag", "Samstag", "Sonntag"
]

if uploaded_files:
    eintraege = []

    for file in uploaded_files:
        try:
            df = pd.read_excel(file, sheet_name="Touren", header=None)
            df = df.iloc[4:]
            df.columns = range(df.shape[1])

            for _, row in df.iterrows():
                kommentar = str(row[15]) if 15 in row and pd.notnull(row[15]) else ""
                datum = pd.to_datetime(row[14], errors='coerce') if 14 in row else None

                # Paar 1: D/E (3/4)
                if 3 in row and pd.notnull(row[3]) and 4 in row and pd.notnull(row[4]):
                    nname = str(row[3])
                    vname = str(row[4])
                    if (
                        "krank" in kommentar.lower()
                        and datum is not None
                        and datum.year >= 2025
                    ):
                        kw = datum.isocalendar().week
                        wochentag = wochentage[datum.weekday()]
                        datum_kw = datum.strftime("%d.%m.%Y") + f" (KW {kw}, {wochentag})"
                        monat_index = datum.month
                        jahr = datum.year
                        monat_name = german_months[monat_index]
                        eintraege.append({
                            "Nachname": nname,
                            "Vorname": vname,
                            "DatumKW": datum_kw,
                            "Kommentar": kommentar,
                            "Monat": f"{monat_index:02d}-{jahr}_{monat_name} {jahr}"
                        })

                # Paar 2: G/H (6/7)
                if 6 in row and pd.notnull(row[6]) and 7 in row and pd.notnull(row[7]):
                    nname = str(row[6])
                    vname = str(row[7])
                    if (
                        "krank" in kommentar.lower()
                        and datum is not None
                        and datum.year >= 2025
                    ):
                        kw = datum.isocalendar().week
                        wochentag = wochentage[datum.weekday()]
                        datum_kw = datum.strftime("%d.%m.%Y") + f" (KW {kw}, {wochentag})"
                        monat_index = datum.month
                        jahr = datum.year
                        monat_name = german_months[monat_index]
                        eintraege.append({
                            "Nachname": nname,
                            "Vorname": vname,
                            "DatumKW": datum_kw,
                            "Kommentar": kommentar,
                            "Monat": f"{monat_index:02d}-{jahr}_{monat_name} {jahr}"
                        })

        except Exception as e:
            st.error(f"Fehler in Datei {file.name}: {e}")

    if eintraege:
        df_gesamt = pd.DataFrame(eintraege)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Monatsblätter
            for monat_key in sorted(df_gesamt["Monat"].unique()):
                df_monat = df_gesamt[df_gesamt["Monat"] == monat_key]
                # Anzahl Kranktage pro Name in diesem Monat berechnen
                kranktage_pro_name = (
                    df_monat.groupby(["Nachname", "Vorname"]).size().to_dict()
                )
                zeilen = []
                for (nach, vor), gruppe in df_monat.groupby(["Nachname", "Vorname"]):
                    kranktage = kranktage_pro_name.get((nach, vor), 0)
                    zeilen.append([f"{vor} {nach} ({kranktage} Tage)", ""])
                    zeilen.append(["Datum", "Kommentar"])
                    for _, r in gruppe.iterrows():
                        zeilen.append([r["DatumKW"], r["Kommentar"]])
                    zeilen.append(["", ""])

                df_sheet = pd.DataFrame(zeilen, columns=["Spalte A", "Spalte B"])
                sheet_name = monat_key.split("_")[1][:31]
                df_sheet.to_excel(writer, index=False, sheet_name=sheet_name)

                sheet = writer.sheets[sheet_name]
                sheet.row_dimensions[1].hidden = True  # Zeile 1 ausblenden

                thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))

                orange_fill = PatternFill("solid", fgColor="ffc000")
                header_fill = PatternFill("solid", fgColor="95b3d7")

                for row in sheet.iter_rows():
                    val = str(row[0].value).strip().lower() if row[0].value else ""
                    is_name_row = (
                        str(row[0].value).strip() != ""
                        and (row[1].value is None or row[1].value == "")
                    )

                    for cell in row:
                        cell.font = Font(name="Calibri", size=11)
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.border = thin

                        if is_name_row:
                            cell.font = Font(bold=True, size=12)
                            cell.fill = orange_fill
                        elif row[0].value == "Datum":
                            cell.font = Font(bold=True)
                            cell.fill = header_fill

                # Autobreite auf alle Spalten
                for col_cells in sheet.columns:
                    max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
                    col_letter = get_column_letter(col_cells[0].column)
                    sheet.column_dimensions[col_letter].width = int(max_len * 1.2) + 2

            # Übersicht-Tab erzeugen
            df_uebersicht = (
                df_gesamt.groupby(["Nachname", "Vorname"])
                .size()
                .reset_index(name="Anzahl Krank-Meldungen")
                .sort_values(by="Anzahl Krank-Meldungen", ascending=False)
            )

            ueberschrift = [["Krank-Meldungen Übersicht ab 2025"], [""], ["Nachname", "Vorname", "Anzahl Krank-Meldungen"]]
            daten = df_uebersicht.values.tolist()
            df_uebersicht_final = pd.DataFrame(ueberschrift + daten, columns=["A", "B", "C"])

            df_uebersicht_final.to_excel(writer, index=False, header=False, sheet_name="Übersicht")
            sheet_ue = writer.sheets["Übersicht"]

            # Merge-Überschrift (NUR EINMAL!)
            sheet_ue.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

            thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
            header_fill = PatternFill("solid", fgColor="95b3d7")
            zebra_1 = PatternFill("solid", fgColor="f7f7f7")
            zebra_2 = PatternFill("solid", fgColor="ddeeff")

            for row_idx, row in enumerate(sheet_ue.iter_rows(), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    # Überschrift
                    if row_idx == 1:
                        cell.font = Font(size=16, bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        # Kein cell.border!
                    # Leerzeile
                    elif row_idx == 2:
                        continue  # Keine Formatierung!
                    # Kopfzeile der Tabelle
                    elif row_idx == 3:
                        cell.font = Font(bold=True)
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin
                    # Datenzeilen
                    elif row_idx > 3:
                        if row_idx % 2 == 0:
                            cell.fill = zebra_1
                        else:
                            cell.fill = zebra_2
                        # Anzahl-Spalte rechtsbündig & fett
                        if col_idx == 3:
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin

            for col in range(1, 4):
                max_len = 18
                if col == 3:
                    max_len = 24
                sheet_ue.column_dimensions[get_column_letter(col)].width = max_len

        st.download_button("Excel-Datei herunterladen", output.getvalue(), file_name="Krank_Monatsauswertung.xlsx")

    else:
        st.warning("Keine gültigen Krank-Meldungen ab 2025 gefunden.")
